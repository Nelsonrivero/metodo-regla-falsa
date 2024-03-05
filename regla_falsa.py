import openpyxl

def regla_falsa(func, a, b, tol=1e-6, max_iter=100):
    """
    Método de la regla falsa para encontrar la raíz de una función en un intervalo dado.

    Parameters:
    - func: Función para la cual se busca la raíz.
    - a, b: Extremos del intervalo inicial.
    - tol: Tolerancia para la convergencia (por defecto, 1e-6).
    - max_iter: Número máximo de iteraciones (por defecto, 100).

    Returns:
    - Aproximación de la raíz, o None si no se alcanza la convergencia en max_iter iteraciones.
    """

    iter_count = 0
    resultados = []

    while iter_count < max_iter:
        fa = func(a)
        fb = func(b)

        if abs(fa) < tol:
            resultados.append((iter_count, a, b, fa, fb, a, fa, None))
            return a, resultados

        if abs(fb) < tol:
            resultados.append((iter_count, a, b, fa, fb, b, fb, None))
            return b, resultados

        # Calcular el siguiente punto usando la regla falsa
        c = (a * fb - b * fa) / (fb - fa)

        fc = func(c)

        if iter_count == 0:
            resultados.append((iter_count, a, b, fa, fb, c, fc, None))
        else:
            resultados.append((iter_count, a, b, fa, fb, c, fc, c - resultados[-1][-3]))

        if abs(fc) < tol:
            return c, resultados

        # Actualizar los extremos del intervalo
        if fa * fc < 0:
            b = c
        else:
            a = c

        iter_count += 1

    # Si no se alcanza la convergencia en max_iter iteraciones
    print("El método de la regla falsa no convergió en {} iteraciones.".format(max_iter))
    return None, resultados

# Solicitar al usuario los extremos del intervalo
a = float(input("Ingrese el extremo izquierdo del intervalo: "))
b = float(input("Ingrese el extremo derecho del intervalo: "))

# Ejemplo de uso:
def ejemplo_funcion(x):
    return x**2-2

# Llamada al método de la regla falsa
raiz, iteraciones = regla_falsa(ejemplo_funcion, a, b)

# Mostrar el resultado
if raiz is not None:
    print("Aproximación de la raíz:", raiz)

    # Crear un archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Iteraciones"

    # Agregar información del intervalo en la primera fila
    ws['A1'] = "Intervalo Izquierdo (a)"
    ws['B1'] = "Intervalo Derecho (b)"
    ws['A2'] = a
    ws['B2'] = b

    # Encabezados de iteraciones
    ws['A4'] = "Iteración"
    ws['B4'] = "a"
    ws['C4'] = "b"
    ws['D4'] = "f(a)"
    ws['E4'] = "f(b)"
    ws['F4'] = "c"
    ws['G4'] = "f(c)"
    ws['H4'] = "c - c_anterior"

    # Llenar datos
    for i, (iter_num, ai, bi, fai, fbi, ci, fci, diff_c) in enumerate(iteraciones, start=5):
        ws.cell(row=i, column=1, value=iter_num)
        ws.cell(row=i, column=2, value=ai)
        ws.cell(row=i, column=3, value=bi)
        ws.cell(row=i, column=4, value=fai)
        ws.cell(row=i, column=5, value=fbi)
        ws.cell(row=i, column=6, value=ci)
        ws.cell(row=i, column=7, value=fci)
        ws.cell(row=i, column=8, value=diff_c)

    # Guardar el archivo Excel
    wb.save("iteraciones_regla_falsa.xlsx")
    print("Se ha creado un archivo Excel con las iteraciones: iteraciones_regla_falsa.xlsx")

else:
    print("No se pudo encontrar una raíz con la tolerancia dada.")
